#!/usr/bin/env python3
# generate_xlsx.py — dipanggil dari PHP via exec()
# Args: [type] [json_input_file] [output_file]
# type: pml | lk

import sys, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def thin_border():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(ws, row, cols, bg='1A2744', fg='FFFFFF', sz=11):
    fill = PatternFill('solid', fgColor=bg)
    font = Font(bold=True, color=fg, size=sz, name='Arial')
    aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = aln
        cell.border = thin_border()

def data_cell(ws, r, c, val, bold=False, color=None, align='left', num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, name='Arial', size=10,
                     color=color if color else '000000')
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin_border()
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def pct_color(v):
    if v >= 50: return '16A34A'
    if v >= 25: return 'D97706'
    return 'E8560A'

if __name__ == '__main__':
    typ      = sys.argv[1]  # 'pml' or 'lk'
    inp_file = sys.argv[2]
    out_file = sys.argv[3]

    with open(inp_file, encoding='utf-8') as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ── PML EXPORT ───────────────────────────────────────────────────────
    if typ == 'pml':
        tanggal = data.get('tanggal', '')
        rows    = data.get('pmlRows', [])

        ws = wb.active
        ws.title = 'Rekap PML'
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20

        # Judul
        ws.merge_cells('A1:N1')
        c = ws['A1']
        c.value = f'REKAP PROGRESS SE2026 PER PML — {tanggal}'
        c.font  = Font(bold=True, size=13, name='Arial', color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='1A2744')
        c.alignment = Alignment(horizontal='center', vertical='center')

        # Header
        headers = [
            'PML', 'Email PCL', 'Nama PCL', 'Total',
            'Dikerjakan', 'Dik%', '+Dik',
            'Diperiksa', 'Per%', '+Per',
            'Approved', 'App%', '+App',
            'Submitted'
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=2, column=ci, value=h)
        header_style(ws, 2, len(headers))

        row_num = 3
        for p in rows:
            pencacah = p.get('pencacah', [])
            for pc in pencacah:
                tot = pc.get('total', 0) or 0
                dik = pc.get('dikerjakan', 0) or 0
                per = pc.get('diperiksa', 0) or 0
                app = pc.get('approved', 0) or 0
                sub = pc.get('submitted', 0) or 0

                dik_pct = round(dik/tot*100, 1) if tot else 0
                per_pct = round(per/tot*100, 1) if tot else 0
                app_pct = round(app/tot*100, 1) if tot else 0

                sdik = pc.get('selisihDikerjakan')
                sper = pc.get('selisihDiperiksa')
                sapp = pc.get('selisihApproved')

                vals = [
                    p.get('pml', ''), pc.get('email', ''), pc.get('nama', ''), tot,
                    dik, dik_pct, sdik,
                    per, per_pct, sper,
                    app, app_pct, sapp,
                    sub
                ]
                for ci, v in enumerate(vals, 1):
                    bold = ci in (1, 3)
                    color = None
                    if ci == 6:  color = pct_color(dik_pct)
                    if ci == 9:  color = pct_color(per_pct)
                    if ci == 12: color = pct_color(app_pct)
                    if ci in (7, 10, 13) and v is not None:
                        color = '16A34A' if v > 0 else ('DC2626' if v < 0 else '6B7280')
                    align = 'right' if ci >= 4 else 'left'
                    cell = data_cell(ws, row_num, ci, v, bold=bold, color=color, align=align)
                    if ci in (6, 9, 12):
                        cell.number_format = '0.0"%"'

                # Alternating row
                if row_num % 2 == 0:
                    for ci in range(1, len(headers)+1):
                        ws.cell(row=row_num, column=ci).fill = PatternFill('solid', fgColor='F7F7F5')

                row_num += 1

        # Lebar kolom
        widths = [20, 28, 22, 8, 10, 8, 7, 10, 8, 7, 10, 8, 7, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:N{row_num-1}'

    # ── LK BEBAN KERJA ───────────────────────────────────────────────────
    elif typ == 'lk':
        tanggal = data.get('tanggal', '')
        rows    = data.get('rows', [])

        ws = wb.active
        ws.title = 'LK Beban Kerja'
        ws.sheet_view.showGridLines = False

        # Header dokumen
        ws.merge_cells('A1:P1')
        c = ws['A1']
        c.value = 'LEMBAR KERJA BEBAN KERJA PETUGAS LAPANGAN SE2026'
        c.font  = Font(bold=True, size=12, name='Arial', color='FFFFFF')
        c.fill  = PatternFill('solid', fgColor='1A2744')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:P2')
        c = ws['A2']
        c.value = f'Kota Kupang — Provinsi Nusa Tenggara Timur | Tanggal: {tanggal}'
        c.font  = Font(size=10, name='Arial', color='5A6478')
        c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 18

        # Header kolom
        headers = [
            'No', 'Nama PML', 'Nama PCL', 'Email PCL',
            'Kecamatan', 'Kode Kec', 'Kode Desa',
            'Target\nFasih', 'Ruta\nTarget', 'Usaha\nTarget',
            'Dikerjakan', 'Dik%',
            'Diperiksa', 'Per%',
            'Approved', 'App%'
        ]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=3, column=ci, value=h)
        header_style(ws, 3, len(headers))
        ws.row_dimensions[3].height = 36

        NAMA_KEC = {
            '5371010': 'ALAK',
            '5371020': 'MAULAFA',
            '5371030': 'OEBOBO',
            '5371031': 'KOTA RAJA',
            '5371040': 'KELAPA LIMA',
            '5371041': 'KOTA LAMA',
        }

        row_num = 4
        for r in rows:
            tot = r.get('total', 0) or 0
            dik = r.get('dikerjakan', 0) or 0
            per = r.get('diperiksa', 0) or 0
            app = r.get('approved', 0) or 0
            dik_pct = round(dik/tot*100, 1) if tot else 0
            per_pct = round(per/tot*100, 1) if tot else 0
            app_pct = round(app/tot*100, 1) if tot else 0
            kode_kec = r.get('kode_kec', '')

            vals = [
                r.get('no', row_num-3),
                r.get('pml', ''),
                r.get('nama', ''),
                r.get('email', ''),
                NAMA_KEC.get(kode_kec, kode_kec),
                kode_kec,
                r.get('kode_desa', ''),
                tot, '', '',  # Ruta & Usaha kosong
                dik, dik_pct,
                per, per_pct,
                app, app_pct
            ]
            for ci, v in enumerate(vals, 1):
                color = None
                if ci == 12: color = pct_color(dik_pct)
                if ci == 14: color = pct_color(per_pct)
                if ci == 16: color = pct_color(app_pct)
                align = 'right' if ci >= 8 else ('center' if ci == 1 else 'left')
                data_cell(ws, row_num, ci, v, color=color, align=align)

            if row_num % 2 == 0:
                for ci in range(1, len(headers)+1):
                    ws.cell(row=row_num, column=ci).fill = PatternFill('solid', fgColor='F7F7F5')

            row_num += 1

        # Keterangan kolom Ruta & Usaha
        ws.cell(row=row_num+1, column=1,
                value='*) Kolom Ruta & Usaha diisi manual — data tidak tersedia di sistem SELARAS')
        ws.cell(row=row_num+1, column=1).font = Font(size=9, color='9E9E9E', italic=True, name='Arial')

        widths = [4, 20, 22, 28, 22, 10, 12, 8, 8, 8, 10, 8, 10, 8, 10, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A4'
        ws.auto_filter.ref = f'A3:P{row_num-1}'

    wb.save(out_file)
    print('OK')